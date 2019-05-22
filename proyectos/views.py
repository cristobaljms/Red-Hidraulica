from django.shortcuts import render
from .models import *
from .utils import *
from materiales.models import Material
from fluidos.models import Fluido
from django.views import generic
from django.contrib import messages
from django.shortcuts import redirect
from django.core.files.storage import FileSystemStorage
from django.core.serializers import serialize
from django.http import HttpResponseRedirect
from django.http import JsonResponse
from django.http import HttpResponse
from io import BytesIO

import numpy as np
from numpy.linalg import inv

from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image, BaseDocTemplate, Frame
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus.tables import Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.lib.pagesizes import A4, landscape, portrait
from reportlab.lib.enums import TA_JUSTIFY, TA_CENTER, TA_LEFT, TA_RIGHT

from openpyxl import Workbook
from openpyxl.styles import Font

import re
import json
import random
import copy

BIN_LIST_2 = ['00', '01', '10', '11']
BIN_LIST_3 = ['000', '001', '010', '011','100', '101', '110', '111']
ITERACION_MAX = 100

def getMatrizBinarios(nindividuos, ntuberias, l):
    matriz = []
    i = 0
    while i < nindividuos:
        if l < 5:
            x = [random.choice(BIN_LIST_2) for i in range(ntuberias)]
        else:
            x = [random.choice(BIN_LIST_3) for i in range(ntuberias)]
        if not(x in matriz):
            matriz.append(x)
            i = i + 1  
    return np.matrix(matriz)

def getMatrizDiametros(matrizBinarios, data_genetico):
    dimension = matrizBinarios.shape
    matriz = np.zeros(dimension)
    for i in range(0, dimension[0]):
        for j in range(0, dimension[1]):
            for dg in data_genetico:
                if str(matrizBinarios[i, j]) == dg['codigo']:
                    matriz[i, j] = dg['diametro']   
    return matriz

def getMatrizCostos(matrizBinarios, data_genetico):
    dimension = matrizBinarios.shape
    matriz = np.zeros(dimension)
    for i in range(0, dimension[0]):
        for j in range(0, dimension[1]):
            for dg in data_genetico:
                if str(matrizBinarios[i, j]) == dg['codigo']:
                    matriz[i, j] = dg['costo']   
    return matriz

def getGeneticData(project_pk, nindividuos):
    diametros = DiametrosGeneticos.objects.filter(proyecto=project_pk).order_by('diametro')
    
    if len(diametros) == 0:
        return "NOT_LOAD_GENETIC_DIAMETER"

    data_genetico = []
    for dg in diametros:
        data_genetico.append({
            "diametro":dg.diametro,
            "costo": dg.costo
        })
        
    l = len(diametros)

    if l <= 4:
        for i in range(l):
            data_genetico[i]['codigo'] = BIN_LIST_2[i]
    else:
        for i in range(l):
            data_genetico[i]['codigo'] = BIN_LIST_3[i]

    if l == 1:
        data_genetico.append({ 'codigo':BIN_LIST_2[1], 'diametro': data_genetico[0]['diametro'], 'costo': data_genetico[0]['costo']})
        data_genetico.append({ 'codigo':BIN_LIST_2[2], 'diametro': data_genetico[0]['diametro'], 'costo': data_genetico[0]['costo']})
        data_genetico.append({ 'codigo':BIN_LIST_2[3], 'diametro': data_genetico[0]['diametro'], 'costo': data_genetico[0]['costo']})
    elif l == 2:
        data_genetico.append({ 'codigo':BIN_LIST_2[2], 'diametro': data_genetico[0]['diametro'], 'costo': data_genetico[0]['costo']})
        data_genetico.append({ 'codigo':BIN_LIST_2[3], 'diametro': data_genetico[1]['diametro'], 'costo': data_genetico[1]['costo']})
    elif l == 3:
        data_genetico.append({ 'codigo':BIN_LIST_2[3], 'diametro': data_genetico[0]['diametro'], 'costo': data_genetico[0]['costo']})
    elif l == 5:
        data_genetico.append({ 'codigo':BIN_LIST_3[5], 'diametro': data_genetico[0]['diametro'], 'costo': data_genetico[0]['costo']})
        data_genetico.append({ 'codigo':BIN_LIST_3[6], 'diametro': data_genetico[1]['diametro'], 'costo': data_genetico[1]['costo']})
        data_genetico.append({ 'codigo':BIN_LIST_3[7], 'diametro': data_genetico[2]['diametro'], 'costo': data_genetico[2]['costo']})
    elif l == 6:
        data_genetico.append({ 'codigo':BIN_LIST_3[6], 'diametro': data_genetico[0]['diametro'], 'costo': data_genetico[0]['costo']})
        data_genetico.append({ 'codigo':BIN_LIST_3[7], 'diametro': data_genetico[1]['diametro'], 'costo': data_genetico[1]['costo']})
    elif l == 7:
        data_genetico.append({ 'codigo':BIN_LIST_3[7], 'diametro': data_genetico[0]['diametro'], 'costo': data_genetico[0]['costo']})
    
    tuberias = Tuberia.objects.filter(proyecto=project_pk)
    K = 0
    for t in tuberias:
        K += t.longitud

    K = getK(K*diametros[0].costo)

    ntuberias = len(tuberias)
    ndiametros = len(data_genetico)
    # matrizBinarios = np.matrix([['11', '10', '10', '11', '10', '00', '10', '01', '00', '11', '00'],
    #                             ['11', '01', '00', '11', '00', '10' ,'00', '01', '00', '01', '00'],
    #                             ['00', '11', '01', '10', '11', '01', '01' ,'10', '11', '00', '11'],
    #                             ['10', '11', '01', '11' ,'00' ,'01', '00' ,'11', '10', '00', '01'],
    #                             ['10', '01', '01', '10', '11', '11', '01', '00', '01', '01', '11']])
    matrizBinarios = getMatrizBinarios(nindividuos, ntuberias, ndiametros)
    matrizDiametros = getMatrizDiametros(matrizBinarios, data_genetico)
    matrizCostos = getMatrizCostos(matrizBinarios, data_genetico)

    return [matrizBinarios, matrizDiametros, matrizCostos, K, data_genetico]

def calculoFO(project_pk, matrizBinarios, matrizDiametros, matrizCostos, projectData, nindividuos):
    result = []
    
    ntuberias = len(projectData['tuberias'])

    qx = 0
    for nodo in projectData['nodos']:
        qx = qx + nodo['demanda']

    Qx = np.zeros(ntuberias) + (qx/ntuberias)

    array_longitud = []
    for t in projectData['tuberias']:
        array_longitud.append(t['longitud'])
    Lx = np.array(array_longitud)

    for i in range(nindividuos):
        calculos = calculosGradiente(1, project_pk, matrizDiametros[i] ,Qx, [], [], [])

        if calculos == "ERROR_MAX_LIMIT_ITERATION":
            return "ERROR_MAX_LIMIT_ITERATION"

        data_maxima = None
        iteracion_maxima = 1
        for c in calculos:
            if (iteracion_maxima <= c['iteracion']):
                iteracion_maxima = c['iteracion']
                data_maxima = c

        # Calculo de Cc
        Cc = 0
        for j in range(ntuberias):
            Cc += matrizCostos[i,j]*Lx[j]

        # Calculo de Cph
        cont = 0
        Cph = 0
        K = getK(Cc)

        ncont = 0
        for n in projectData['nodos']:
            Pmin = 10
            Pnodo = data_maxima['H'][cont] - n['cota']
            AP = Pmin - Pnodo
            if AP <= 0:
                AP = 0
            else:
                C = np.power(2.71828,AP)
                AP = C
                ncont += 1
            cont += 1
            if Cph <= AP:
                Cph = AP
        Cph *= K

        #Calculo de Cv
        Cv = 0
        tcont = 0
        for t in range(ntuberias):
            Vmin = 0.3
            Vtubo = data_maxima['tabla'][t]['V']
            AV = Vmin - Vtubo
            if AV <= 0:
                AV = 0
            else:
                Cv += AV
                tcont += 1
        Cv*=K

        FO = Cc + Cph + Cv
        # logging.info("FO:{} Cc:{} Cph:{} Cv:{}".format(FO, np.round(Cc,1), np.round(Cph,1), np.round(Cv,1)))
        # logging.info("binarios {}".format(concatArr(matrizBinarios[i].tolist()[0])              ))
        result.append({
            'FO':FO,
            'binarios': concatArr(matrizBinarios[i].tolist()[0]),
            'ncont': ncont,
            'tcont': tcont
        })
    return bubbleSort(result, 'FO')

def seleccion(FO, Nc, B):
    Pmax = B/Nc
    Pmin = (2-B)/Nc

    arrProbabilidad = []

    for i in range(1, Nc+1):
        Pi = Pmin + ( Pmax - Pmin ) * ((Nc - i)/(Nc - 1))
        PiN = Pi * Nc
        if PiN >= 1.5:
            arrProbabilidad.append(FO[i-1])
            arrProbabilidad.append(FO[i-1])
        elif PiN >= 0.5:
            arrProbabilidad.append(FO[i-1])

    arrPadresCruzamiento = []

    for i in arrProbabilidad:
        arrPadresCruzamiento.append(i['binarios'])

    arrIndividuos=[i+1 for i in range(Nc)]

    flagA = True
    while(flagA):
        auxArr = copy.deepcopy(arrIndividuos)
        arrIndexPadres = []
        while(len(auxArr) > 1):
            a = random.choice(auxArr)
            b = random.choice(auxArr)
            if (a != b and (a + 1) != b and (a-1) != b):
                auxArr.remove(a)
                auxArr.remove(b)
                arrIndexPadres.append({'a':a, 'b':b})    
            if(len(auxArr) == 2 and ((auxArr[0] + 1) == auxArr[1] or (auxArr[1] + 1) == auxArr[0])):
                break
            elif( len(auxArr) == 1 ):
                flagA = False
                arrIndexPadres.append({'a':auxArr[0], 'b':-1})
            elif(len(auxArr) == 0):
                flagA = False
    return [arrIndexPadres, arrProbabilidad]

def cruzamiento(data):
    arrIndexPadres = data[0]
    arrProbabilidad = data[1]
    print(arrIndexPadres)
    print(arrProbabilidad)
    arrHijosCruzamiento = []
    for aIP in arrIndexPadres:

        if aIP['b'] == -1:
            print("===============================")
            print("impar")
            print(arrProbabilidad[aIP['a']-1])
            arrHijosCruzamiento.append(arrProbabilidad[aIP['a']-1])
            print("===============================")
        else:
            
            binA = arrProbabilidad[aIP['a']-1]['binarios']
            binB = arrProbabilidad[aIP['b']-1]['binarios']
            print("===============================")
            print("padres")
            print("a: {}".format(binA))
            print("b: {}".format(binB))
            arrbinA = [binA[c] for c in range(len(binA))]
            arrbinB = [binB[c] for c in range(len(binB))]

            a = random.randint(1, len(binA) -1)
            b = random.randint(1, len(binA) -1)
            if a > b:
                a, b = b, a
            print("ramd a {}".format(a))
            print("ramd b {}".format(b))

            for i in range(a ,b-1):
                arrbinA[i], arrbinB[i] = arrbinB[i], arrbinA[i]

            arrProbabilidad[aIP['a']-1]['binarios'] = concatArr(arrbinA)
            arrProbabilidad[aIP['b']-1]['binarios'] = concatArr(arrbinB)
            print("hijos")
            print("a: {}".format(arrProbabilidad[aIP['a']-1]['binarios']))
            print("b: {}".format(arrProbabilidad[aIP['b']-1]['binarios']))
            print("===============================")
            arrHijosCruzamiento.append(arrProbabilidad[aIP['a']-1])
            arrHijosCruzamiento.append(arrProbabilidad[aIP['b']-1])
    print(arrHijosCruzamiento)
    return arrHijosCruzamiento

def mutacion(hijosCruzamiento):
    Lc = len(hijosCruzamiento[0]['binarios'])
    Pm = 1/Lc

    listado_mutacion = []
    for d in hijosCruzamiento:
        a = random.uniform(0.0001, 1)

        if a >= Pm:
            listado_mutacion.append(d)
        else:
            a = random.randint(0, Lc-1)
            arrBinarios = [hijosCruzamiento[0]['binarios'][c] for c in range(Lc)]

            if arrBinarios[a] == '0':
                arrBinarios[a] = '1'
            else:
                arrBinarios[a] = '0' 
   
            hijosCruzamiento[0]['binarios'] = concatArr(arrBinarios)  

    return hijosCruzamiento

from celery import task, shared_task, current_task
import logging

@shared_task
def calculosGenetico(project_pk):
    
    try:
        dataGenetica = DatosGeneticos.objects.get(proyecto=project_pk)
    except DatosGeneticos.DoesNotExist:
        return "NO_GENETIC_DATA_LOAD"

    nindividuos = dataGenetica.nindividuos
    
    pos = int(np.round(nindividuos*0.8,0))-1

    npoblacion = dataGenetica.npoblacion
    B = dataGenetica.beta

    geneticData = getGeneticData(project_pk, nindividuos)
    if geneticData == "NOT_LOAD_GENETIC_DIAMETER":
        return "NOT_LOAD_GENETIC_DIAMETER"

    matrizBinarios = geneticData[0]
    matrizDiametros = geneticData[1]
    matrizCostos = geneticData[2]

    K = geneticData[3]
    data_genetico = geneticData[4]
    projectData = getProjectData(project_pk)
    matrizBinariosFromMutacion = []
    result = []
    # print("Matriz binario inicial aleatoria")
    # print(matrizBinarios)

    for i in range(npoblacion):
        logging.info("poblacion {}".format(i))
        process_percent = int(100 * float(i) / float(npoblacion))
        current_task.update_state(state='PROGRESS', meta={'current': i,'total':npoblacion, 'percent':process_percent})
        if len(matrizBinariosFromMutacion) == 0:
            resultado_FO = calculoFO(project_pk, matrizBinarios, matrizDiametros, matrizCostos, projectData, nindividuos)
        else:
            matrizDiametros = getMatrizDiametros(matrizBinariosFromMutacion, data_genetico)
            matrizCostos = getMatrizCostos(matrizBinariosFromMutacion, data_genetico)
            resultado_FO = calculoFO(project_pk, matrizBinariosFromMutacion, matrizDiametros, matrizCostos, projectData, nindividuos)
        
        if resultado_FO == "ERROR_MAX_LIMIT_ITERATION":
            return "ERROR_MAX_LIMIT_ITERATION"
        # logging.info(resultado_FO)
        # # with open('your_file.txt', 'w') as f:
        # #     for item in resultado_FO:
        # #         f.write("%s\n" % item)

        # print(i,resultado_FO)
        # resultado_FO['npoblacion'] = i
        result.append([i,resultado_FO])
        # print("FO")
        # print(resultado_FO)
        resultado_seleccion = seleccion(resultado_FO, nindividuos, B)
        # print("Resultado Calculo de seleccion")
        # print(resultado_seleccion)
        resultado_cruzamiento = cruzamiento(resultado_seleccion)
        # print("Resultado Calculo de cruzamiento")
        # print(resultado_cruzamiento)
        resultado_mutacion = mutacion(resultado_cruzamiento)
        # print("Resultado Calculo de mutacion")
        # print(resultado_mutacion)
        arrBinarios = [rm['binarios'] for rm in resultado_mutacion]
        # print("arrBinarios")
        # print(arrBinarios)
        matrizBinariosFromMutacion = handleArrMutacionToMatrizBinarios(arrBinarios, 4)
        # print("Matriz binarios desde la mutacion")
        # print(matrizBinariosFromMutacion)
        if(validate_result_fo(resultado_FO, pos, K)):
            break
    return result

def GeneticoToPDFView(request):
    data = request.POST['data']

    pdf_buffer = BytesIO()
    doc = SimpleDocTemplate(pdf_buffer, pagesize=portrait(A4))
    Story = []

    ps_head = ParagraphStyle('titulo',alignment = TA_CENTER, fontSize = 14, fontName="Times-Roman")
    ps_subtitle = ParagraphStyle('titulo',alignment = TA_JUSTIFY, fontSize = 14, fontName="Times-Roman")
    ps_iteracion = ParagraphStyle('titulo',alignment = TA_JUSTIFY, fontSize = 12, fontName="Times-Roman")
    ps_tabla = ParagraphStyle('titulo',alignment = TA_JUSTIFY, fontSize = 8, fontName="Times-Roman")

    text = "<b>ALGORITMO GENETICO</b>"
    p = Paragraph(text, ps_head)
    Story.append(p)
    Story.append(Spacer(1,0.5*inch))

    titles = [
        Paragraph('<b>FO</b>', ps_tabla),
        Paragraph('<b>Binario</b>', ps_tabla),
    ]
    
    text = "<b>1. Mejores 10 individuos de cada generación</b>"
    p = Paragraph(text, ps_subtitle)
    Story.append(p)
    Story.append(Spacer(1,0.2*inch))

    poblacion_cont = 1 
    for poblacion in json.loads(data):
        Story.append(Spacer(1,0.1*inch))
        text = "<p>Poblacion {}</p>".format(poblacion_cont)
        p = Paragraph(text, ps_iteracion)
        Story.append(p)
        Story.append(Spacer(1,0.2*inch))

        table_formatted = [titles]
        individuo_cont = 0
        for i in poblacion[1]:
            if individuo_cont == 10:
                break
            binarios=i['binarios']
            FO=np.round(i['FO'],0)
            row = [ FO, binarios ]
            table_formatted.append(row)
            individuo_cont += 1

        t=Table(table_formatted, (100,300))

        t.setStyle(TableStyle([
            ('BACKGROUND',(0,0),(2,0),'#878787'),
            ('INNERGRID',(0,0),(2,0), 0.25, colors.gray),
            ('BOX',(0,0),(2,0), 0.25, colors.gray)
        ]))

        Story.append(t)
        poblacion_cont += 1

    
    #punto 3
    Story.append(Spacer(1,0.2*inch))
    text = "<b>2. 10 mejores individuos de todo el proceso</b>"
    p = Paragraph(text, ps_subtitle)
    Story.append(p)
    Story.append(Spacer(1,0.2*inch))

    todo_array = []
    for poblacion in json.loads(data):
        todo_array += poblacion[1]
    
    todo_array = bubbleSort(todo_array, 'FO')

    table_formatted = [titles]
 
    for i in range(10):
        binarios=todo_array[i]['binarios']
        FO=np.round(todo_array[i]['FO'],0)
        row = [ FO, binarios ]
        table_formatted.append(row)

    t=Table(table_formatted, (100,300))

    t.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(2,0),'#878787'),
        ('INNERGRID',(0,0),(2,0), 0.25, colors.gray),
        ('BOX',(0,0),(2,0), 0.25, colors.gray)
    ]))

    Story.append(t)


    # punto 4
    Story.append(Spacer(1,0.2*inch))
    text = "<b>3. Mejor individuo de cada generación</b>"
    p = Paragraph(text, ps_subtitle)
    Story.append(p)
    Story.append(Spacer(1,0.2*inch))

    titles = [
        Paragraph('<b>Poblacion</b>', ps_tabla),
        Paragraph('<b>FO</b>', ps_tabla),
    ]

    poblacion_cont = 1
    table_formatted = [titles]
    for poblacion in json.loads(data):

        FO=np.round(poblacion[1][0]['FO'],0)
        row = [poblacion_cont, FO ]
        table_formatted.append(row)
        poblacion_cont += 1 

    t=Table(table_formatted, (50,100))

    t.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(2,0),'#878787'),
        ('INNERGRID',(0,0),(2,0), 0.25, colors.gray),
        ('BOX',(0,0),(2,0), 0.25, colors.gray)
    ]))
    
    Story.append(t)

    #punto 5
    Story.append(Spacer(1,0.2*inch))
    text = "<b>4. Promedio y desviación estandar de cada generación</b>"
    p = Paragraph(text, ps_subtitle)
    Story.append(p)
    Story.append(Spacer(1,0.2*inch))
    
    titles = [
        Paragraph('<b>Poblacion</b>', ps_tabla),
        Paragraph('<b>Promedio</b>', ps_tabla),
        Paragraph('<b>Desviación</b>', ps_tabla),
    ]
    table_formatted = [titles]
    poblacion_cont = 1
    for poblacion in json.loads(data):
        promedio = 0
        n = len(poblacion[1])
        for i in poblacion[1]:
            promedio += i['FO']
        
        promedio = promedio / n
        desvEstandar = 0
        for i in poblacion[1]:
            desvEstandar += np.power((i['FO']-promedio),2)
        desvEstandar = np.power((desvEstandar/(n-1)),0.5)
        
        row = [poblacion_cont, np.round(promedio,0), np.round(desvEstandar,0) ]
        table_formatted.append(row)

        t=Table(table_formatted, (50,100,100))
        poblacion_cont += 1

    t.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(3,0),'#878787'),
        ('INNERGRID',(0,0),(3,0), 0.25, colors.gray),
        ('BOX',(0,0),(3,0), 0.25, colors.gray)
    ]))
    
    Story.append(t)

    # punto 6
    Story.append(Spacer(1,0.2*inch))
    text = "<b>5. Número y promedio de nodos que no cumplen las restricciones de presión</b>"
    p = Paragraph(text, ps_subtitle)
    Story.append(p)
    Story.append(Spacer(1,0.2*inch))
    
    titles = [
        Paragraph('<b>Individuo</b>', ps_tabla),
        Paragraph('<b>N°</b>', ps_tabla),
        Paragraph('<b>Población</b>', ps_tabla),
    ]
    poblacion_cont = 1
    for poblacion in json.loads(data):
        Story.append(Spacer(1,0.1*inch))
        text = "<p>Poblacion {}</p>".format(poblacion_cont)
        p = Paragraph(text, ps_iteracion)
        Story.append(p)
        Story.append(Spacer(1,0.2*inch))
        poblacion_cont += 1
        individuo_cont = 1
        table_formatted = [titles]

        nodos_promedio = 0
        for i in poblacion[1]:
            row = [individuo_cont, i['ncont'],poblacion[0]+1]
            table_formatted.append(row)
            individuo_cont += 1
            nodos_promedio += i['ncont']
        

        t=Table(table_formatted, (60,40,50))
        t.setStyle(TableStyle([
            ('BACKGROUND',(0,0),(3,0),'#878787'),
            ('INNERGRID',(0,0),(3,0), 0.25, colors.gray),
            ('BOX',(0,0),(3,0), 0.25, colors.gray)
        ]))
        Story.append(t)

        text = "<p><b>Nodos que no cumplen:</b> {}</p>".format(nodos_promedio)
        p = Paragraph(text, ps_iteracion)
        Story.append(p)
        Story.append(Spacer(1,0.2*inch))

        nodos_promedio = nodos_promedio / len(poblacion[1])

        text = "<p><b>Promedio:</b> {}</p>".format(nodos_promedio)
        p = Paragraph(text, ps_iteracion)
        Story.append(p)
        Story.append(Spacer(1,0.2*inch))
    
    Story.append(Spacer(1,0.2*inch))
    text = "<b>6. Número y promedio de tuberias que no cumplen las restricciones</b>"
    p = Paragraph(text, ps_subtitle)
    Story.append(p)
    Story.append(Spacer(1,0.2*inch))
    
    titles = [
        Paragraph('<b>Individuo</b>', ps_tabla),
        Paragraph('<b>N°</b>', ps_tabla),
        Paragraph('<b>Población</b>', ps_tabla),
    ]
    poblacion_cont = 1
    for poblacion in json.loads(data):
        Story.append(Spacer(1,0.1*inch))
        text = "<p>Poblacion {}</p>".format(poblacion_cont)
        p = Paragraph(text, ps_iteracion)
        Story.append(p)
        Story.append(Spacer(1,0.2*inch))
        poblacion_cont += 1
        individuo_cont = 1
        table_formatted = [titles]

        tuberias_promedio = 0
        for i in poblacion[1]:
            row = [individuo_cont, i['tcont'], poblacion[0]+1]
            table_formatted.append(row)
            individuo_cont += 1
            tuberias_promedio += i['tcont']

        t=Table(table_formatted, (60,40,50))
        t.setStyle(TableStyle([
            ('BACKGROUND',(0,0),(3,0),'#878787'),
            ('INNERGRID',(0,0),(3,0), 0.25, colors.gray),
            ('BOX',(0,0),(3,0), 0.25, colors.gray)
        ]))
        Story.append(t)

        text = "<p><b>Tuberias que no cumplen:</b> {}</p>".format(tuberias_promedio)
        p = Paragraph(text, ps_iteracion)
        Story.append(p)
        Story.append(Spacer(1,0.2*inch))

        tuberias_promedio = tuberias_promedio / len(poblacion[1])

        text = "<p><b>Promedio:</b> {}</p>".format(tuberias_promedio)
        p = Paragraph(text, ps_iteracion)
        Story.append(p)
        Story.append(Spacer(1,0.2*inch))

    # Punto 8

    Story.append(Spacer(1,0.2*inch))
    text = "<b>7. Número y promedio de individuos que no cumplen restricciones</b>"
    p = Paragraph(text, ps_subtitle)
    Story.append(p)
    Story.append(Spacer(1,0.2*inch))

    titles = [
        Paragraph('<b>Individuo</b>', ps_tabla),
        Paragraph('<b>N° tuberias</b>', ps_tabla),
        Paragraph('<b>N° Nodos</b>', ps_tabla),
        Paragraph('<b>Status</b>', ps_tabla),
    ]

    poblacion_cont = 1
    for poblacion in json.loads(data):
        Story.append(Spacer(1,0.1*inch))
        text = "<p>Poblacion {}</p>".format(poblacion_cont)
        p = Paragraph(text, ps_iteracion)
        Story.append(p)
        Story.append(Spacer(1,0.2*inch))
        poblacion_cont += 1

        individuo_cont = 1
        table_formatted = [titles]
        individuos_cumplen_cont = len(poblacion[1])
        for i in poblacion[1]:
            status = 'No cumple'
            if (i['tcont'] == 0 and i['ncont'] == 0):
                status = 'Cumple'
                individuos_cumplen_cont -= 1
            row = [individuo_cont, i['tcont'], i['ncont'], status]
            table_formatted.append(row)
            individuo_cont += 1
            
        t=Table(table_formatted, (60,60,60,60))
        t.setStyle(TableStyle([
            ('BACKGROUND',(0,0),(4,0),'#878787'),
            ('INNERGRID',(0,0),(4,0), 0.25, colors.gray),
            ('BOX',(0,0),(4,0), 0.25, colors.gray)
        ]))
        Story.append(t)

        text = "<p><b>Individuos que no cumplen:</b> {}</p>".format(individuos_cumplen_cont)
        p = Paragraph(text, ps_iteracion)
        Story.append(p)
        Story.append(Spacer(1,0.1*inch))
            
        text = "<p><b>Promedio: </b>{}</p>".format(individuos_cumplen_cont/len(poblacion))
        p = Paragraph(text, ps_iteracion)
        Story.append(p)
        Story.append(Spacer(1,0.3*inch))

    # punto 9
    Story.append(Spacer(1,0.2*inch))
    text = "<b>8. Mejores 3 individuos de la ultima iteración</b>"
    p = Paragraph(text, ps_subtitle)
    Story.append(p)
    Story.append(Spacer(1,0.2*inch))

    data = json.loads(data)
    poblacion = data[1][-1]

    titles = [
        Paragraph('<b>FO</b>', ps_tabla),
        Paragraph('<b>Binario</b>', ps_tabla),
    ]
    
    table_formatted = [titles]
    for i in range(3):
        binarios=poblacion[i]['binarios']
        FO=np.round(poblacion[i]['FO'],0)
        row = [ FO, binarios ]
        table_formatted.append(row)

    t=Table(table_formatted, (100,300))

    t.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(2,0),'#878787'),
        ('INNERGRID',(0,0),(2,0), 0.25, colors.gray),
        ('BOX',(0,0),(2,0), 0.25, colors.gray)
    ]))

    Story.append(t)

    doc.build(Story)

    pdf_value = pdf_buffer.getvalue()
    pdf_buffer.close()
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="gradientmethod.pdf"'
    response.write(pdf_value)
    return response

class GeneticView(generic.View):
    template_name = "sections/calculos/genetico.html"

    def get(self, request, *args, **kwargs):
        task_id = calculosGenetico.delay(kwargs['pk'])
        context = {
            'project_pk': kwargs['pk'],
            'task_id':task_id.id
        }
        return render(request, self.template_name, context)

class ProyectoAdminView(generic.CreateView):
    template_name = "sections/proyectos/show.html"

    def get(self, request, *args, **kwargs):
        proyecto = Proyecto.objects.get(pk=kwargs['pk'])
        nodos = Nodo.objects.filter(proyecto=proyecto).order_by('orden')
        tuberias = Tuberia.objects.filter(proyecto=proyecto).order_by('orden')
        reservorios = Reservorio.objects.filter(proyecto=proyecto)
        diametros = DiametrosGeneticos.objects.filter(proyecto=proyecto).order_by("diametro")
        sreservorios = json.loads(serialize("json", reservorios))
        snodos = json.loads(serialize("json", nodos))
        datagenetica = None
        try:
            datagenetica = DatosGeneticos.objects.get(proyecto=proyecto)
        except DatosGeneticos.DoesNotExist:
            datagenetica = { 
                'nindividuos':' ',
                'npoblacion' :' ',
                'porcentaje_replicacion' :' ',
                'porcentaje_mutacion' :' ',
                'porcentaje_cruzami':' ' 
            }
            
        rarray = []

        for sr in sreservorios:
            fields = sr['fields']
            fields['pktype'] = 'r'+str(sr['pk'])
            rarray.append(fields)

        for sn in snodos:
            fields = sn['fields']
            fields['pktype'] = 'n'+str(sn['pk'])
            rarray.append(fields)

        torden = len(tuberias)
        norden = len(nodos)
        
        context = {
            'proyecto': proyecto,
            'nodos':nodos,
            'tuberias':tuberias,
            'diametros':diametros,
            'datagenetica':datagenetica,
            'norden': norden + 1,
            'torden': torden + 1,
            'reservorios': reservorios,
            'opciones_tuberia': rarray,
            'active_tab': kwargs['active_tab'], 
        }
        
        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        tipo = request.POST.get('tipo')
        id_proyecto = request.POST.get('id_proyecto')

        if (tipo == 'nodo'):
            active_tab = 'n'
            numero = request.POST.get('numero')
            demanda = request.POST.get('demanda')
            cota = request.POST.get('cota')
            orden = request.POST.get('orden')
            x_position = request.POST.get('x_position')
            y_position = request.POST.get('y_position')
            proyecto = Proyecto.objects.get(pk=id_proyecto)
            nodo = Nodo(proyecto=proyecto, numero=numero, cota=cota, demanda=demanda, orden=orden, x_position=x_position, y_position=y_position)
            nodo.save()
            messages.add_message(request, messages.SUCCESS, 'Nodo creado con exito')
            return redirect('proyecto_administrar', id_proyecto, active_tab)

        elif (tipo == 'tuberia'):
            active_tab = 't'
            numero = request.POST.get('numero')
            longitud = request.POST.get('longitud')
            diametro = request.POST.get('diametro')
            km = request.POST.get('km')
            orden = request.POST.get('orden')
            start = request.POST.get('start')
            end = request.POST.get('end')

            if start == '0' or end == '0':
                messages.add_message(request, messages.ERROR, 'Debe ingresar el nodo incial y el nodo final')
                return redirect('proyecto_administrar', id_proyecto, active_tab)

            if start == end:
                messages.add_message(request, messages.ERROR, 'El nodo de inicio no puede ser igual al nodo final')
                return redirect('proyecto_administrar', id_proyecto, active_tab)

            if(re.match('n', start)):
                patron = re.compile('n')
                nstart = Nodo.objects.get(pk = int(patron.split(start)[1]))  
            else:
                patron = re.compile('r')
                nstart = Reservorio.objects.get(pk = int(patron.split(start)[1]))

            if(re.match('n', end)):
                patron = re.compile('n')
                nend = Nodo.objects.get(pk = int(patron.split(end)[1]))  
            else:
                patron = re.compile('r')
                nend = Reservorio.objects.get(pk = int(patron.split(end)[1]))

            proyecto = Proyecto.objects.get(pk=id_proyecto)
            tuberia = Tuberia(proyecto=proyecto, numero=numero, orden=orden, longitud=longitud, diametro=diametro, km=km, start=nstart.numero, end = nend.numero)
            tuberia.save()
            messages.add_message(request, messages.SUCCESS, 'Tuberia creada con exito')
            return redirect('proyecto_administrar', id_proyecto, active_tab)

        elif (tipo == 'genetico'):
            active_tab = 'g'
            
            l = DiametrosGeneticos.objects.filter(proyecto=id_proyecto).count()
            if l == 8:
                messages.add_message(request, messages.ERROR, 'Solo se pueden cargar 8 diametros como maximo')
                return redirect('proyecto_administrar', id_proyecto, active_tab)

            proyecto = Proyecto.objects.get(pk=id_proyecto)
            diametro = request.POST.get('diametro')
            costo = request.POST.get('costo')
            dg = DiametrosGeneticos(proyecto=proyecto, diametro=diametro, costo=costo)
            dg.save()
            messages.add_message(request, messages.SUCCESS, 'Diametro creado con exito')
            return redirect('proyecto_administrar', id_proyecto, active_tab)
        
        elif (tipo == 'datagenetico'):
            active_tab = 'g'
            nindividuos = request.POST.get('nindividuos')
            npoblacion = request.POST.get('npoblacion')
            # preplicacion = request.POST.get('preplicacion')
            # pmutacion = request.POST.get('pmutacion')
            # pcruzami = request.POST.get('pcruzami')
            preplicacion = 0
            pmutacion = 0
            pcruzami = 0
            beta = request.POST.get('beta')
            try:
                datagentica = DatosGeneticos.objects.get(proyecto=id_proyecto)
                datagentica.nindividuos = nindividuos
                datagentica.npoblacion = npoblacion
                datagentica.porcentaje_replicacion = preplicacion
                datagentica.porcentaje_mutacion = pmutacion
                datagentica.porcentaje_cruzami = pcruzami
                datagentica.beta = beta
                datagentica.save()
                messages.add_message(request, messages.SUCCESS, 'data guardada con exito')
                return redirect('proyecto_administrar', id_proyecto, active_tab)

            except DatosGeneticos.DoesNotExist:
                proyecto = Proyecto.objects.get(pk=id_proyecto)
                datagentica = DatosGeneticos(proyecto = proyecto, nindividuos = nindividuos, npoblacion = npoblacion,porcentaje_replicacion = preplicacion,porcentaje_mutacion = pmutacion, porcentaje_cruzami = pcruzami, beta=beta)
                datagentica.save()
                messages.add_message(request, messages.SUCCESS, 'data guardada con exito')
                return redirect('proyecto_administrar', id_proyecto, active_tab)

        else:
            active_tab = 'r'
            numero = request.POST.get('numero')
            z = request.POST.get('z')
            x_position = request.POST.get('x_position')
            y_position = request.POST.get('y_position')
            
            proyecto = Proyecto.objects.get(pk=id_proyecto)
            reservorio = Reservorio(numero=numero, z=z, proyecto=proyecto, y_position=y_position, x_position=x_position)
            reservorio.save()
            messages.add_message(request, messages.SUCCESS, 'Reservorio creado con exito')
            return redirect('proyecto_administrar', id_proyecto, active_tab)

class ProyectosListView(generic.ListView):
    model = Proyecto
    template_name = 'sections/proyectos/index.html'

class ProyectosCreateView(generic.CreateView):
    template_name = "sections/proyectos/create.html"

    def get(self, request, *args, **kwargs):
        fluidos = Fluido.objects.all()
        material = Material.objects.all()
        context = { 'fluidos':fluidos, 'material':material }
        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        nombre = request.POST.get('nombre')
        fluido = request.POST.get('fluido')
        material = request.POST.get('material')

        if len(nombre) < 2:
            messages.add_message(request, messages.ERROR, 'El nombre del proyecto debe ser mayor a 2 digitos')
            return redirect('proyectos_crear')

        if fluido == '0':
            messages.add_message(request, messages.ERROR, 'El fluido no es valido')
            return redirect('proyectos_crear')
        
        if material == '0':
            messages.add_message(request, messages.ERROR, 'El material no es valido')
            return redirect('proyectos_crear')

        if Proyecto.objects.filter(nombre=nombre).exists():
            messages.add_message(request, messages.ERROR, 'Este nombre ya se encuentra registrado.')
            return render(request, self.template_name)
        else:
            f = Fluido.objects.get(pk=fluido)
            m = Material.objects.get(pk=material)
            p = Proyecto(nombre=nombre, fluido=f, material=m)
            p.save()
        
        messages.add_message(request, messages.SUCCESS, 'Proyecto creado con exito')
        return redirect('proyectos')
        
class ProyectosUpdateView(generic.View):
    template_name = "sections/proyectos/edit.html"

    def get(self, request, *args, **kwargs):
        proyecto = Proyecto.objects.get(pk=kwargs['pk'])
        fluidos = Fluido.objects.all()
        materiales = Material.objects.all()
        context = {
            'proyecto': proyecto, 
            'fluidos': fluidos,
            'materiales':materiales
        }
        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        nombre = request.POST.get('nombre')
        fluido = request.POST.get('fluido')
        material = request.POST.get('material')
        id_proyecto = request.POST.get('id')

        try:
            p = Proyecto.objects.get(pk=id_proyecto)
        except Proyecto.DoesNotExist:
            messages.add_message(request, messages.ERROR, 'No existe el proyecto')
            return redirect('proyectos')
        
        if len(nombre) < 2:
            messages.add_message(request, messages.ERROR, 'El nombre del proyecto debe ser mayor a 2 digitos')
            return redirect('proyectos_editar', pk=id_proyecto)

        if fluido == '0':
            messages.add_message(request, messages.ERROR, 'El fluido no es valido')
            return redirect('proyectos_editar', pk=id_proyecto)
        
        if material == '0':
            messages.add_message(request, messages.ERROR, 'El material no es valido')
            return redirect('proyectos_editar', pk=id_proyecto)

        f = Fluido.objects.get(pk=fluido)
        m = Material.objects.get(pk=material)
        p.nombre = nombre
        p.fluido = f
        p.material = m
        p.save()

        messages.add_message(request, messages.SUCCESS, 'Proyecto editado con exito')
        return redirect('proyectos')

class ProyectoDeleteView(generic.DeleteView):
    template_name = "sections/proyectos/delete.html"
    def get(self, request, *args, **kwargs):
        proyecto = Proyecto.objects.get(pk=kwargs['pk'])
        context = {
            'proyecto': proyecto
        }
        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        Proyecto.objects.filter(pk=kwargs['pk']).delete()
        messages.add_message(request, messages.SUCCESS, 'Proyecto eliminado')
        return redirect('proyectos')

class TuberiaUpdateView(generic.View):

    def post(self, request, *args, **kwargs):
        active_tab = 't'
        id_tuberia = kwargs['pk']
        id_proyecto = request.POST.get('id_proyecto')
        numero = request.POST.get('numero')
        longitud = request.POST.get('longitud')
        diametro = request.POST.get('diametro')
        km = request.POST.get('km')
        orden = request.POST.get('orden')
        start = request.POST.get('mstart')
        end = request.POST.get('mend')

        if start == '0' or end == '0':
            messages.add_message(request, messages.ERROR, 'Debe ingresar el nodo incial y el nodo final')
            return redirect('proyecto_administrar', id_proyecto, active_tab)

        if start == end:
            messages.add_message(request, messages.ERROR, 'El nodo de inicio no puede ser igual al nodo final')
            return redirect('proyecto_administrar', id_proyecto, active_tab)

        if(re.match('n', start)):
            patron = re.compile('n')
            nstart = Nodo.objects.get(pk = int(patron.split(start)[1]))  
        else:
            patron = re.compile('r')
            nstart = Reservorio.objects.get(pk = int(patron.split(start)[1]))

        if(re.match('n', end)):
            patron = re.compile('n')
            nend = Nodo.objects.get(pk = int(patron.split(end)[1]))  
        else:
            patron = re.compile('r')
            nend = Reservorio.objects.get(pk = int(patron.split(end)[1]))

        tuberia = Tuberia.objects.get(pk=id_tuberia)
        tuberia.numero=numero
        tuberia.orden=orden
        tuberia.longitud=longitud
        tuberia.diametro=diametro
        tuberia.km=km
        tuberia.start=nstart.numero
        tuberia.end = nend.numero
        tuberia.save()

        messages.add_message(request, messages.SUCCESS, 'Tuberia actualizada con exito')
        return redirect('proyecto_administrar', id_proyecto, active_tab)

class NodoUpdateView(generic.View):

    def post(self, request, *args, **kwargs):
        active_tab = 'n'
        numero = request.POST.get('numero')
        demanda = request.POST.get('demanda')
        cota = request.POST.get('cota')
        orden = request.POST.get('orden')
        x_position = request.POST.get('x_position')
        y_position = request.POST.get('y_position')
        id_proyecto = request.POST.get('id_proyecto')

        nodo = Nodo.objects.get(pk=kwargs['pk'])
        nodo.numero=numero
        nodo.cota=cota
        nodo.demanda=demanda
        nodo.orden=orden
        nodo.x_position=x_position
        nodo.y_position=y_position
        nodo.save()

        messages.add_message(request, messages.SUCCESS, 'Nodo actualizado con exito')
        return redirect('proyecto_administrar', id_proyecto, active_tab)

class DiametroGeneticoUpdateView(generic.View):

    def post(self, request, *args, **kwargs):
        active_tab = 'g'
        diametro = request.POST.get('diametro')
        costo = request.POST.get('costo')
        id_proyecto = request.POST.get('id_proyecto')
        dg = DiametrosGeneticos.objects.get(pk=kwargs['pk'])
        dg.diametro=diametro
        dg.costo=costo
        dg.save()

        messages.add_message(request, messages.SUCCESS, 'Diametro actualizado con exito')
        return redirect('proyecto_administrar', id_proyecto, active_tab)

def borrarTuberia(request, pk):
    Tuberia.objects.filter(pk=pk).delete()
    return HttpResponseRedirect(request.META.get('HTTP_REFERER'))

def borrarNodo(request, pk):
    Nodo.objects.filter(pk=pk).delete()
    return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
    
def borrarReservorio(request, pk):
    Reservorio.objects.filter(pk=pk).delete()
    return HttpResponseRedirect(request.META.get('HTTP_REFERER'))

def borrarDiametroGenetico(request, pk):
    DiametrosGeneticos.objects.filter(pk=pk).delete()
    return HttpResponseRedirect(request.META.get('HTTP_REFERER'))

def getProjectData(pk):
    tuberias = json.loads(serialize("json", Tuberia.objects.filter(proyecto=pk).order_by('orden')))
    nodos = json.loads(serialize("json", Nodo.objects.filter(proyecto=pk).order_by('orden')))
    reservorios = json.loads(serialize("json", Reservorio.objects.filter(proyecto=pk)))
    
    tarray = []
    for t in tuberias:
        tarray.append(t['fields'])

    narray = []
    for n in nodos:
        narray.append(n['fields'])

    rarray = []
    for r in reservorios:
        rarray.append(r['fields'])

    context = {
        'tuberias' : tarray,
        'nodos' : narray,
        'reservorios' : rarray,
    }
    return context

def obtenerProyectoDatos(request, pk):
    return JsonResponse(getProjectData(pk), safe=False)

def calculosGradiente(iteracion, pk, Dx, Qx, H, A12, response):
    
    if (iteracion > ITERACION_MAX):
        return "ERROR_MAX_LIMIT_ITERATION"

    data = getProjectData(pk)
    proyecto = Proyecto.objects.get(pk=pk)
    iteracionRow = { "iteracion": iteracion }

    ntuberias = len(data['tuberias'])
    nnodos = len(data['nodos'])
    nreservorios = len(data['reservorios'])
    
    # Creacion del arreglo de longitud
    array_longitud = []
    for t in data['tuberias']:
        array_longitud.append(t['longitud'])
    Lx = np.array(array_longitud)
    

    # Creacion del arreglo de diametro
    if len(Dx) == 0:
        array_diametro = []
        for t in data['tuberias']:
            array_diametro.append(t['diametro'])
        Dx = np.array(array_diametro)
    

    # Creacion del arreglo de km
    array_km = []
    for t in data['tuberias']:
        array_km.append(t['km'])
    Km = np.array(array_km)


    # Guardamos Qx en una variable auxiliar para luego saber que fila de A12
    # debe ser multiplicada por -1 en caso de que una fila de Qx sea negativa
    aux_qx = np.zeros(ntuberias) + Qx


    # Convertimos a Qx en positivo
    for i in range(0,ntuberias):
        if (Qx[i] < 0):
            Qx[i] = Qx[i] * -1

    # Calculamos el Area
    A = (np.pi*np.power(Dx,2))/4

    # Calculamos la Velocidad
    V = Qx/A

    # Armamos el arrego de Ks 
    Ks   = np.zeros(ntuberias) + proyecto.material.ks

    # Calculamos Re
    Re   = np.zeros(ntuberias) + V*Dx/proyecto.fluido.valor_viscocidad
    Re = np.round(Re, 0)

    # Calculamos el coheficiente de friccion
    f = []
    for i in range(0,ntuberias):
        f.append(f_calculo(Re[i],Ks[i]/Dx[i]))

    # Calculamos hf
    hf   = np.zeros(ntuberias) + f*(Lx/Dx)*(np.power(V,2)/(2*9.81))

    # Calculamos hm
    hm   = np.zeros(ntuberias) + Km * (np.power(V,2)/(2*9.81))

    # Calculamos la suma de hf y hm
    hfhm = np.zeros(ntuberias) + (hf + hm)

    # Calculamos alfa
    a    = np.zeros(ntuberias) + (hfhm / np.power(Qx, 2))

    # Calculamos la multiplicacion de alfa por Qx
    af   = np.zeros(ntuberias) + (a * Qx)

    # Creamos un json on todos estos datos para enviarlos luego a las Vistas y reportes
    table = TableFormatter(ntuberias,data['tuberias'], Qx, Lx, Dx, A,V,Re, f,hf,Km,hm,hfhm,a, af)
    iteracionRow['tabla'] = table
    
    """
    Ahora comenzamos a crear y calcular las matrices
    """
    # Matriz de conectividad

    if (len(A12) == 0):
        j = 0
        for tuberia in data['tuberias']:
            a = np.zeros(nnodos).astype(int)
            for i in range(0, nnodos):
                if(tuberia['start'] == data['nodos'][i]['numero']):
                    a[i] = -1
            for i in range(0, nnodos):
                if(tuberia['end'] == data['nodos'][i]['numero']):
                    a[i] = 1
            if (aux_qx[j] < 0):
                a = a * -1
            j = j + 1
            A12.append(a)
        A12 = np.matrix(A12)
    else:
        for j in range(0, ntuberias):
            if (aux_qx[j] < 0):
                A12[j] = A12[j] * -1
    # Matriz traspuesta de A12
    A21 = A12.transpose()

    # MATRIZ TOPOLOGICA A10
    A10 = []
    for tuberia in data['tuberias']:
        a = np.zeros(nreservorios).astype(int)
        for i in range(0, nreservorios):
            if(tuberia['start'] == data['reservorios'][i]['numero'] or tuberia['end'] == data['reservorios'][i]['numero']):
                a[i] = -1
        A10.append(a)
    A10 = np.matrix(A10)

    # MATRIZ DIAGONAL A11 
    A11 = np.zeros((ntuberias, ntuberias))
    for i in range(0, len(af)):
        A11[i][i] = af[i]
    
    # Arreglo alturas de reservorios
    H0 = []
    for reservorio in data['reservorios']:
        H0.append(reservorio['z'])
    
    H0 = np.array(H0)

    # Arreglo caudal de salida
    q = []
    for nodo in data['nodos']:
        q.append(nodo['demanda'])
    
    q = np.array(q)
    q = np.reshape(q, (nnodos,1))
   
    # Matriz diagonal del 2 y matriz identidad
    N = np.zeros((ntuberias, ntuberias)).astype(int)
    I = np.zeros((ntuberias, ntuberias)).astype(int)
    for i in range(0, ntuberias):
        N[i][i] = 2
        I[i][i] = 1
    
   
    ##### CALCULAMOS LAS H #####

    # ([N][A11])^-1
    step1 = inv(N*A11)

    # [A21]([N][A11])^-1
    step2 = A21*step1

    # ([A21]([N][A11])^-1)*([A12])
    step3 = step2*A12

    # -(([A21]([N][A11])^-1)*([A12])^-1
    step4 = inv(step3) * -1

    Qx = np.reshape(Qx, ntuberias)

    # [A11][Q]+[A10][H0]
    step5 = Qx.dot(A11) + A10.dot(H0)
    step5 = np.reshape(step5, (ntuberias,1))

    # [A21]([N][A11])^-1*([A11][Q]+[A10][H0])
    step6 = step2.dot(step5)

    Qx = np.reshape(Qx, (ntuberias,1))

    # A21*Qx
    step7 = A21.dot(Qx)

    # A21*Qx-q
    step8 = step7 - q 
    step9 = step6 - step8

    # step10 son las H ya calculadas
    step10 = step4.dot(step9)
    
    # Las guardamos para enviarlas luego a los reportes
    iteracionRow['H'] = np.squeeze(np.asarray(np.round(step10,4))).tolist()

    #### CALCULAMOS LAS Q ####

    Qstep1 = inv(N*A11)
    Qstep2 = Qstep1*A11
    Qstep3 = I - Qstep2
    Qx = np.reshape(Qx, (ntuberias,1))
    Qstep4 = Qstep3.dot(Qx)
    Qstep5 = A10 * H0
    Qstep6 = A12 * step10
    Qstep7 = Qstep6 + Qstep5
    Qstep8 = Qstep1.dot(Qstep7)

    # En Qstep9 se encuentra el rsultado de todas las Qx
    Qstep9 = Qstep4 - Qstep8
    
    # Guardamos Qx para enviarlo luego a los reportes y vistas
    iteracionRow['Qx'] = np.squeeze(np.asarray(np.round(Qstep9, 4))).tolist()
   
    # Culminamos la iteracion y subimos este valor en 1
    iteracion = iteracion + 1

    # Si H esta vacia entonces es la primera iteracion y no calculamos el error
    if(len(H) > 0):
        # Calculamos el error
        error = np.absolute(H-step10)

        # Lo almacenamos y lo aadimos a la respuesta
        iteracionRow['error'] = np.squeeze(np.asarray(np.round(error,4))).tolist()
        response.append(iteracionRow)

        # Validamos el error, si deuvelve True entonces llamamos recursivamente esta funcion
        # Con los nuevos parametros
        if (validateError(error)):
            Qstep9 = np.squeeze(np.asarray(Qstep9))
            return calculosGradiente(iteracion, pk, Dx, Qstep9, step10, A12, response)
        else:
            # Sino retornamos y finaliza el calculo
            return response
    else:
        # Esto solo ocurrira en la primera iteracion, donde no hay que calcular el error
        response.append(iteracionRow)
        Qstep9 = np.squeeze(np.asarray(Qstep9))
        return calculosGradiente(iteracion, pk, Dx, Qstep9, step10, A12, response)

class GradienteView(generic.View):
    template_name = "sections/calculos/gradiente.html"

    def get(self, request, *args, **kwargs):
        data = getProjectData(kwargs['pk'])
        ntuberias = len(data['tuberias'])
        qx = 0
        for nodo in data['nodos']:
            qx = qx + nodo['demanda']
        Qx = np.zeros(ntuberias) + (qx/ntuberias)

        calculos = calculosGradiente(1, kwargs['pk'], [],  Qx, [], [], [])

        context = {
            'data': calculos,
            'project_pk': kwargs['pk']
        }
        #return JsonResponse(context, safe=False)
        return render(request, self.template_name, context)

def GradienteToPDFView(request, pk):
    data = getProjectData(pk)
    ntuberias = len(data['tuberias'])
    
    qx = 0
    for nodo in data['nodos']:
        qx = qx + nodo['demanda']

    Qx = np.zeros(ntuberias) + (qx/ntuberias)

    calculos = calculosGradiente(1, pk, [], Qx, [], [], [])

    if calculos == "ERROR_MAX_LIMIT_ITERATION":
        active_tab = 'i'
        messages.add_message(request, messages.ERROR, 'Se ha superado el limite de iteraciones que es {}, se sugiere:\n - Revisar que los datos cargados estan correctos'.format(ITERACION_MAX))
        return redirect('proyecto_administrar', pk, active_tab)

    pdf_buffer = BytesIO()
    doc = SimpleDocTemplate(pdf_buffer, pagesize=landscape(A4))
    Story = []

    ps_head = ParagraphStyle('titulo',alignment = TA_CENTER, fontSize = 14, fontName="Times-Roman")
    ps_iteracion = ParagraphStyle('titulo',alignment = TA_JUSTIFY, fontSize = 12, fontName="Times-Roman")
    ps_tabla = ParagraphStyle('titulo',alignment = TA_JUSTIFY, fontSize = 8, fontName="Times-Roman")

    text = "<b>METODO DE LA GRADIENTE</b>"
    p = Paragraph(text, ps_head)
    Story.append(p)
    Story.append(Spacer(1,0.5*inch))

    titles = [
        Paragraph('<b>Tuberia</b>', ps_tabla),
        Paragraph('<b>Caudal</b>', ps_tabla),
        Paragraph('<b>Longitud</b>', ps_tabla),
        Paragraph('<b>Diametro</b>', ps_tabla),
        Paragraph('<b>Area</b>', ps_tabla),
        Paragraph('<b>Velocidad</b>', ps_tabla),
        Paragraph('<b>f</b>', ps_tabla),
        Paragraph('<b>Km</b>', ps_tabla),
        Paragraph('<b>hf+hm</b>', ps_tabla),
        Paragraph('<b>a</b>', ps_tabla),
        Paragraph('<b>a*Qx</b>', ps_tabla)
    ]
    
    for iteracion in calculos:
        text = "<b>Iteracion {}</b>".format(iteracion['iteracion'])
        p = Paragraph(text, ps_iteracion)
        Story.append(p)
        Story.append(Spacer(1,0.2*inch))

        table_formatted = [titles]
        for i in iteracion['tabla']:
            tuberia = i['tuberia']
            Qx = i['Qx']
            Lx = i['Lx']
            Dx = i['Dx']
            A = i['A']
            V = i['V']
            Re = i['Re']
            f = i['f']
            hf = i['hf']
            Km = i['Km']
            hm = i['hm']
            hfhm = i['hfhm']
            a = i['a']
            af = i['af']
            row = [ tuberia, Qx, Lx, Dx, A, V, f, Km, hfhm, a, af ]
            table_formatted.append(row)

        t=Table(table_formatted, (60,40,60,60,60,60,50,40,100,100,80))

        t.setStyle(TableStyle([
            ('BACKGROUND',(0,0),(13,0),'#878787'),
            ('INNERGRID',(0,0),(13,0), 0.25, colors.gray),
            ('BOX',(0,0),(13,0), 0.25, colors.gray)
        ]))

        Story.append(t)
        Story.append(Spacer(1,0.2*inch))

        THtitles = [ 
            Paragraph('H', ps_tabla)
        ]
        table_formatted = [THtitles]
        for value in iteracion['H']:
            table_formatted.append([value])

        t=Table(table_formatted, (40))
        t.setStyle(TableStyle([
            ('BACKGROUND',(0,0),(1,0),colors.lightgrey),
            ('INNERGRID',(0,0),(1,0), 0.25, colors.gray),
            ('BOX',(0,0),(1,0), 0.25, colors.gray)
        ]))
        Story.append(t)
        Story.append(Spacer(1,0.2*inch))

        TQxtitles = [ 
            Paragraph('Qx', ps_tabla)
        ]

        table_formatted = [TQxtitles]
        for value in iteracion['Qx']:
            table_formatted.append([value])

        t=Table(table_formatted, (40))
        t.setStyle(TableStyle([
            ('BACKGROUND',(0,0),(1,0),colors.lightgrey),
            ('INNERGRID',(0,0),(1,0), 0.25, colors.gray),
            ('BOX',(0,0),(1,0), 0.25, colors.gray)
        ]))
        Story.append(t)
        Story.append(Spacer(1,0.2*inch))


        TErrortitles = [ 
            Paragraph('Error', ps_tabla)
        ]

        table_formatted = [TErrortitles]

        if 'error' in iteracion:
            for value in iteracion['error']:
                table_formatted.append([value])

            t=Table(table_formatted, (40))
            t.setStyle(TableStyle([
                ('BACKGROUND',(0,0),(1,0),colors.lightgrey),
                ('INNERGRID',(0,0),(1,0), 0.25, colors.gray),
                ('BOX',(0,0),(1,0), 0.25, colors.gray)
            ]))
            Story.append(t)
            Story.append(Spacer(1,0.2*inch))

        #table_formatted = [titles]
        frameCount = 2
        frames = []
        #construct a frame for each column
        for frame in range(frameCount):
            column = Frame(100, 50, 50, 50)
            frames.append(column)

        #Story.append(frames)
        Story.append(Spacer(1,0.2*inch))

    proyecto = Proyecto.objects.get(pk=pk)

    doc.build(Story)
    pdf_value = pdf_buffer.getvalue()
    pdf_buffer.close()
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="gradientmethod.pdf"'
    response.write(pdf_value)
    return response

def GradienteToExcelView(request, pk):
    data = getProjectData(pk)
    ntuberias = len(data['tuberias'])
    project = Proyecto.objects.get(pk=pk)

    qx = 0
    for nodo in data['nodos']:
        qx = qx + nodo['demanda']

    Qx = np.zeros(ntuberias) + (qx/ntuberias)
    
    calculos = calculosGradiente(1, pk, [], Qx, [], [], [])
    
    if calculos == "ERROR_MAX_LIMIT_ITERATION":
        active_tab = 'i'
        messages.add_message(request, messages.ERROR, 'Se ha superado el limite de iteraciones que es {}, se sugiere:\n - Revisar que los datos cargados estan correctos'.format(ITERACION_MAX))
        return redirect('proyecto_administrar', pk, active_tab)

    wb = Workbook()
    ws = wb.active
    ws.title = "data"
    cont = 1
    ws.merge_cells('A1:D1')
    ws['A'+str(cont)] = 'PROPIEDADES DEL RESERVORIO'
    ws['A'+str(cont)].font = Font(size=12, bold=True)

    cont = 2
    ws['A'+str(cont)] = 'ID'
    ws['A'+str(cont)].font = Font(size=10, bold=True)
    ws['B'+str(cont)] = 'LGH (Zf)'
    ws['B'+str(cont)].font = Font(size=10, bold=True)
    ws['C'+str(cont)] = 'X'
    ws['C'+str(cont)].font = Font(size=10, bold=True)
    ws['D'+str(cont)] = 'Y'
    ws['D'+str(cont)].font = Font(size=10, bold=True)

    cont = 3
    for r in data['reservorios']:
        ws.cell(row=cont,column=1).value = r['numero']
        ws.cell(row=cont,column=2).value = r['z']
        ws.cell(row=cont,column=3).value = r['x_position']
        ws.cell(row=cont,column=4).value = r['y_position']
        cont = cont + 1

    cont = cont + 1
    ws.merge_cells('A'+str(cont)+':D'+str(cont))
    ws['A'+str(cont)] = 'PROPIEDADES DE LOS NODOS'
    ws['A'+str(cont)].font = Font(size=12, bold=True)

    cont = cont + 1
    ws['A'+str(cont)] = 'ID'
    ws['A'+str(cont)].font = Font(size=10, bold=True)
    ws['B'+str(cont)] = 'X'
    ws['B'+str(cont)].font = Font(size=10, bold=True)
    ws['C'+str(cont)] = 'Y'
    ws['C'+str(cont)].font = Font(size=10, bold=True)
    ws['D'+str(cont)] = 'Zn (msnm)'
    ws['D'+str(cont)].font = Font(size=10, bold=True)
    ws['E'+str(cont)] = 'DEMANDA m3/seg (Qs)'
    ws['E'+str(cont)].font = Font(size=10, bold=True)
    
    cont = cont + 1
    for n in data['nodos']:
        ws.cell(row=cont,column=1).value = n['numero']
        ws.cell(row=cont,column=2).value = n['x_position']
        ws.cell(row=cont,column=3).value = n['y_position']
        ws.cell(row=cont,column=4).value = n['cota']
        ws.cell(row=cont,column=5).value = n['demanda']
        cont = cont + 1

    cont = cont + 1
    ws.merge_cells('A'+str(cont)+':D'+str(cont))
    ws['A'+str(cont)] = 'PROPIEDADES DE LAS TUBERIAS'
    ws['A'+str(cont)].font = Font(size=12, bold=True)

    cont = cont + 1
    ws['A'+str(cont)] = 'TUBERIA'
    ws['A'+str(cont)].font = Font(size=10, bold=True)
    ws['B'+str(cont)] = 'NODO INICIAL'
    ws['B'+str(cont)].font = Font(size=10, bold=True)
    ws.column_dimensions['B'].width = 12
    ws['C'+str(cont)] = 'NODO FINAL'
    ws['C'+str(cont)].font = Font(size=10, bold=True)
    ws.column_dimensions['C'].width = 12
    ws['D'+str(cont)] = 'DIAMETRO (m)'
    ws['D'+str(cont)].font = Font(size=10, bold=True)
    ws.column_dimensions['D'].width = 13
    ws['E'+str(cont)] = 'LONGITUD (m)'
    ws['E'+str(cont)].font = Font(size=10, bold=True)
    ws.column_dimensions['E'].width = 13
    ws['F'+str(cont)] = 'COE_MENORES'
    ws['F'+str(cont)].font = Font(size=10, bold=True)

    cont = cont + 1
    for n in data['tuberias']:
        ws.cell(row=cont,column=1).value = n['numero']
        ws.cell(row=cont,column=2).value = n['start']
        ws.cell(row=cont,column=3).value = n['end']
        ws.cell(row=cont,column=4).value = n['diametro']
        ws.cell(row=cont,column=5).value = n['longitud']
        ws.cell(row=cont,column=6).value = n['km']
        cont = cont + 1

    cont = cont + 1
    ws.merge_cells('A'+str(cont)+':D'+str(cont))
    ws['A'+str(cont)] = 'PROPIEDADES FISICAS'
    ws['A'+str(cont)].font = Font(size=12, bold=True)

    cont = cont + 1
    ws['A'+str(cont)] = 'GRAVEDAD (g)'
    ws['A'+str(cont)].font = Font(size=10, bold=True)
    ws['B'+str(cont)] = 9.8
    ws['C'+str(cont)] = 'm/seg^2'

    cont = cont + 1
    ws['A'+str(cont)] = 'TEMPERATURA'
    ws['A'+str(cont)].font = Font(size=10, bold=True)
    ws['B'+str(cont)] =  project.material.ks
    ws['C'+str(cont)] = 'º C'

    cont = cont + 1
    ws['A'+str(cont)] = 'VISCOSIDAD CINEMÁTICA'
    ws['A'+str(cont)].font = Font(size=10, bold=True)
    ws['B'+str(cont)] = project.fluido.valor_viscocidad
    ws['C'+str(cont)] = 'm^2/seg'

    ws1 = wb.create_sheet(title="resultados")

    cont = 1
    ws1.merge_cells('A'+str(cont)+':D'+str(cont))
    ws1['A'+str(cont)] = 'RESULTADOS DE LOS NODOS'
    ws1['A'+str(cont)].font = Font(size=12, bold=True)

    cont = 2
    ws1['A'+str(cont)] = 'ID'
    ws1['A'+str(cont)].font = Font(size=10, bold=True)
    ws1['B'+str(cont)] = 'Zn (msnm)'
    ws1['B'+str(cont)].font = Font(size=10, bold=True)
    ws1['C'+str(cont)] = 'DEMANDA'
    ws1['C'+str(cont)].font = Font(size=10, bold=True)
    ws1['D'+str(cont)] = 'LGH'
    ws1['D'+str(cont)].font = Font(size=10, bold=True)
    ws1['E'+str(cont)] = 'Presion'
    ws1['E'+str(cont)].font = Font(size=10, bold=True)

    data_maxima = None
    iteracion_maxima = 1
    for c in calculos:
        if (iteracion_maxima <= c['iteracion']):
            data_maxima = c

    cont = cont + 1
    i = 0
    for n in data['nodos']:
        ws1.cell(row=cont,column=1).value = n['numero']
        ws1.cell(row=cont,column=2).value = n['cota']
        ws1.cell(row=cont,column=3).value = n['demanda']
        ws1.cell(row=cont,column=4).value = data_maxima['H'][i]
        i = i + 1
        cont = cont + 1

    cont = cont + 1
    ws1.merge_cells('A'+str(cont)+':D'+str(cont))
    ws1['A'+str(cont)] = 'RESULTADOS DE LAS TUBERIAS'
    ws1['A'+str(cont)].font = Font(size=12, bold=True)

    cont = cont + 1
    ws1['A'+str(cont)] = 'TUBERIA'
    ws1['A'+str(cont)].font = Font(size=10, bold=True)
    ws1['B'+str(cont)] = 'NODO INICIAL'
    ws1['B'+str(cont)].font = Font(size=10, bold=True)
    ws1.column_dimensions['B'].width = 12
    ws1['C'+str(cont)] = 'NODO FINAL'
    ws1['C'+str(cont)].font = Font(size=10, bold=True)
    ws1.column_dimensions['c'].width = 12
    ws1['D'+str(cont)] = 'DIAMETRO (m)'
    ws1['D'+str(cont)].font = Font(size=10, bold=True)
    ws1.column_dimensions['D'].width = 13
    ws1['E'+str(cont)] = 'LONGITUD (m)'
    ws1['E'+str(cont)].font = Font(size=10, bold=True)
    ws1.column_dimensions['E'].width = 13
    ws1['F'+str(cont)] = 'COE_MENORES'
    ws1['F'+str(cont)].font = Font(size=10, bold=True)
    ws1.column_dimensions['F'].width = 13
    ws1['G'+str(cont)] = 'CAUDAL'
    ws1['G'+str(cont)].font = Font(size=10, bold=True)
    ws1['H'+str(cont)] = 'VELOCIDAD'
    ws1['H'+str(cont)].font = Font(size=10, bold=True)
    ws1.column_dimensions['H'].width = 12
    ws1['I'+str(cont)] = 'REYNOLDS'
    ws1['I'+str(cont)].font = Font(size=10, bold=True)
    ws1.column_dimensions['I'].width = 12
    ws1['J'+str(cont)] = 'f'
    ws1['J'+str(cont)].font = Font(size=10, bold=True)
    ws1['K'+str(cont)] = 'hm'
    ws1['K'+str(cont)].font = Font(size=10, bold=True)
    ws1['L'+str(cont)] = 'PERDIDAS\nMENORES'
    ws1['L'+str(cont)].font = Font(size=10, bold=True)
    ws1.column_dimensions['L'].width = 12
    ws1['M'+str(cont)] = 'PERDIDAS\nTOTALES'
    ws1['M'+str(cont)].font = Font(size=10, bold=True)
    ws1.column_dimensions['M'].width = 12
    ws1.row_dimensions[cont].height = 30

    cont = cont + 1
    i = 0
    for n in data['tuberias']:
        ws1.cell(row=cont,column=1).value = n['numero']
        ws1.cell(row=cont,column=2).value = n['start']
        ws1.cell(row=cont,column=3).value = n['end']
        ws1.cell(row=cont,column=4).value = n['diametro']
        ws1.cell(row=cont,column=5).value = n['longitud']
        ws1.cell(row=cont,column=6).value = n['km']
        ws1.cell(row=cont,column=7).value = data_maxima['Qx'][i]
        ws1.cell(row=cont,column=8).value = data_maxima['tabla'][i]['V']
        ws1.cell(row=cont,column=9).value = data_maxima['tabla'][i]['Re']
        ws1.cell(row=cont,column=10).value = data_maxima['tabla'][i]['f']
        ws1.cell(row=cont,column=11).value = data_maxima['tabla'][i]['hm']
        ws1.cell(row=cont,column=12).value = data_maxima['tabla'][i]['hf']
        ws1.cell(row=cont,column=13).value = data_maxima['tabla'][i]['hfhm']
        i = i + 1
        cont = cont + 1

    nombre_archivo ="ReportGradient.xlsx"
    response = HttpResponse(content_type="application/ms-excel") 
    contenido = "attachment; filename={0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    wb.save(response)
    return response
